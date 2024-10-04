import time
from datetime import datetime
from openpyxl import load_workbook, Workbook

from input_person import InputPerson
from output_person import OutputPerson

from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from webdriver_manager.chrome import ChromeDriverManager
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC


def get_chrome_driver():
    """Inicializa e retorna uma instância do ChromeDriver."""

    service = Service(ChromeDriverManager().install())

    chrome_options = Options()
    chrome_options.add_argument("--headless=new")

    return webdriver.Chrome(service=service, options=chrome_options)


def fill_input_field(driver, field_id, value):
    """Preenche os inputs do formulário."""

    input_field = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.ID, field_id))
    )
    input_field.send_keys(value)


def handle_hcaptcha(driver):
    """Lida com a resolução do hCaptcha."""

    WebDriverWait(driver, 10).until(
        EC.frame_to_be_available_and_switch_to_it(
            (By.XPATH, "//iframe[contains(@src,'hcaptcha') and contains(@src,'checkbox')]")
        )
    )
    hcaptcha = WebDriverWait(driver, 10).until(
        EC.element_to_be_clickable((By.XPATH, "/html/body"))
    )
    hcaptcha.click()
    
    time.sleep(5)
    
    driver.switch_to.default_content()


def handle_error(driver, sheet_output, input_person, error_xpath, error_message):
    """Verifica e trata erros na página."""

    error_element = driver.find_elements(By.XPATH, error_xpath)
    if error_element:
        sheet_output.append([
            input_person.formatted_cpf(), '', input_person.formatted_birth_date(), '', '', error_message
        ])
        return True
    return False


def extract_output_person(driver):
    """Extrai informações da página de resultados e retorna um objeto OutputPerson."""

    return OutputPerson(
        cpf=driver.find_element(By.XPATH, '//*[@id="mainComp"]/div[2]/p/span[1]/b').text,
        name=driver.find_element(By.XPATH, '//*[@id="mainComp"]/div[2]/p/span[2]/b').text,
        birth_date=driver.find_element(By.XPATH, '//*[@id="mainComp"]/div[2]/p/span[3]/b').text,
        registration_status=driver.find_element(By.XPATH, '//*[@id="mainComp"]/div[2]/p/span[4]/b').text,
        registration_date=driver.find_element(By.XPATH, '//*[@id="mainComp"]/div[2]/p/span[5]/b').text,
        message='CPF Válido'
    )


def process_person(driver, input_person, sheet_output):
    """Processa os dados de uma pessoa, preenchendo o formulário e salvando os resultados."""

    cpf = input_person.formatted_cpf()
    birth_date = input_person.formatted_birth_date()

    if cpf is None or birth_date is None:
        return

    fill_input_field(driver, 'txtCPF', cpf)
    fill_input_field(driver, 'txtDataNascimento', birth_date)

    time.sleep(5)
    handle_hcaptcha(driver)
    
    btn = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.ID, 'id_submit'))
    )
    btn.send_keys(Keys.ENTER)
    time.sleep(5)

    if handle_error(driver, sheet_output, input_person, '//*[@id="content-core"]/div/div/div[1]/span/h4/b', 'Data de Nascimento Inválida'):
        return

    if handle_error(driver, sheet_output, input_person, '//*[@id="content-core"]/div/div/div[1]/span/h4', 'CPF Inválido'):
        return

    output_person = extract_output_person(driver)

    sheet_output.append([
        output_person.cpf, output_person.name, output_person.birth_date,
        output_person.registration_status, output_person.registration_date, output_person.message
    ])


def save_report(wb_output):
    """Salva o relatório gerado em um arquivo Excel."""

    current_date = datetime.now().strftime('%Y%m%d')
    current_time = datetime.now().strftime('%H%M%S')
    wb_output.save(f'relatorios/RelatorioReceitafederal_{current_date}_{current_time}.xlsx')


def main():
    driver = get_chrome_driver()
    driver.get('https://servicos.receita.fazenda.gov.br/servicos/cpf/consultasituacao/consultapublica.asp')

    wb_input = load_workbook('cpfDataNascimento.xlsx', read_only=True)
    sheet_input = wb_input['Plan1']

    wb_output = Workbook()
    sheet_output = wb_output.active
    sheet_output.title = 'Relatório Receita Federal'
    sheet_output.append(['CPF', 'Nome', 'Data de Nascimento', 'Situação Cadastral', 'Data da Inscrição', 'Mensagem'])

    print("Precessando...")
    for row in sheet_input.iter_rows(min_row=2, values_only=True):
        input_person = InputPerson(row)
        process_person(driver, input_person, sheet_output)
        link = driver.find_element(By.XPATH, '//*[@id="idRodape"]/p[4]/a')
        link.click()
        time.sleep(5)

    print("Processamento finalizado!")
    
    print("Salvando relatório...")
    save_report(wb_output)
    
    driver.quit()
    wb_input.close()
    wb_output.close()

    print("Relatório gerado com sucesso!")


if __name__ == '__main__':
    main()
